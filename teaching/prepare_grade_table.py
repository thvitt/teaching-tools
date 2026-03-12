from csv import DictReader
from itertools import chain
from pathlib import Path
from typing import TYPE_CHECKING, Any, Optional, cast, overload, override

from more_itertools import first
from openpyxl.cell import Cell
from openpyxl.comments.comments import Comment
from openpyxl.styles.colors import Color
from openpyxl.styles.fills import PatternFill

if TYPE_CHECKING:
    from openpyxl.cell import _CellGetValue  # pyright: ignore[reportPrivateUsage]

import logging

import questionary
from attr import dataclass
from openpyxl import Workbook, load_workbook
from openpyxl.styles.fonts import Font
from openpyxl.styles.named_styles import NamedStyle
from openpyxl.worksheet.worksheet import Worksheet
from rich import get_console
from rich.logging import RichHandler
from rich.table import Table
from typer import Typer
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule

logger = logging.getLogger(__name__)

app = Typer()


def column_letter(column: int) -> str:
    """
    Calculates the letter code for a column number.

    Examples:
    >>> column_letter(1)
    'A'
    >>> column_letter(2)
    'B'
    >>> column_letter(28)
    'AB'
    """
    if column < 1:
        raise ValueError(column)

    unhandled = column
    result: list[str] = []
    while unhandled:
        result.append(chr(unhandled % 26 + ord("A") - 1))
        unhandled = unhandled // 26
    return "".join(reversed(result))


class Bookmarks(dict[str, int]):
    """
    Convenience class for accessing indexes in the excel sheet.
    """

    @override
    def __getitem__(self, key: str, /):  # pyright: ignore[reportIncompatibleMethodOverride]
        if key.endswith("_") and key[:-1] in self:
            return column_letter(super().__getitem__(key[:-1]))
        else:
            return super().__getitem__(key)

    def __getattr__(self, name: str) -> int:
        return self[name]  # pyright: ignore[reportReturnType]

    def expand(self, fmt: str) -> str:
        return fmt.format_map(self)

    def using(self, **kwargs: int) -> "Bookmarks":
        copy = self.__class__(**self)
        copy.update(kwargs)
        return copy


class InvalidForm(ValueError): ...


class TaskAssessment:
    name: str
    given_name: str
    full_name: str
    grade: float
    id: int | None = None
    max_grade: float | None = None

    @overload
    def __init__(self, *, full_name: str, grade: float, max_grade: float) -> None: ...

    @overload
    def __init__(
        self, *, name: str, given_name: str, id: int, grade: float
    ) -> None: ...

    def __init__(
        self,
        *,
        name: str = "",
        given_name: str = "",
        full_name: str = "",
        id: int | None = None,
        grade: float = 0.0,
        max_grade: float | None = None,
    ) -> None:
        if name:
            self.name = name
            self.given_name = given_name
            self.full_name = f"{given_name} {name}"
        else:
            self.full_name = full_name
            self.name = self.full_name.split()[-1]
            self.given_name = " ".join(self.full_name.split()[:-1])
        self.id = id
        self.grade = grade
        self.max_grade = max_grade

    @override
    def __str__(self) -> str:
        return f"{self.full_name} ({self.id}): {self.grade}/{self.max_grade or '?'}"

    @override
    def __repr__(self) -> str:
        return f"{self.__class__.__name__}(full_name={self.full_name!r}, id={self.id!r}, grade={self.grade!r}, max_grade={self.max_grade!r})"


@dataclass
class ExamAssassment:
    name: str
    given_name: str
    id: int | None
    exam_id: int
    exam: str = ""

    @property
    def full_name(self):
        return f"{self.given_name} {self.name}"


class GradeSheet:
    workbook: Workbook
    gradesheet: Worksheet
    grades_mapping: Worksheet

    students: dict[str, ExamAssassment]
    tasks: dict[str, dict[str, TaskAssessment]]
    task_abbr: dict[str, str]
    exam_tasks: list[str]

    def __init__(self, files: list[Path]) -> None:
        self.workbook = Workbook()
        _gs = self.workbook.active
        assert _gs is not None
        self.gradesheet = _gs
        self.gradesheet.title = "Bewertung"
        self.grades_mapping = self._create_grade_mapping()

        self.students = {}
        self.tasks = {}
        self.task_abbr = {}
        self.exam_tasks = []

        for file in files:
            if file.suffix == ".xlsx":
                try:
                    self.students.update(
                        {
                            student.full_name: student
                            for student in self.parse_his_sheet(file)
                        }
                    )
                except Exception as e:
                    logger.error("Failed to read HIS sheet %s: %s", file, e)
            elif file.suffix == ".csv":
                ass = parse_wuecampus_sheet(file)
                for task_name, assassments in ass.items():
                    self.tasks[task_name] = {
                        student.full_name: student for student in assassments
                    }

    def merge_student_lists(self):
        for task_name, task in self.tasks.items():
            for student, assassment in task.items():
                if student not in self.students:
                    logger.warning(
                        "Unangemeldet: %s (aus Aufgabe %s)", student, task_name
                    )
                    self.students[student] = ExamAssassment(
                        assassment.name, assassment.given_name, assassment.id, 0
                    )

    def ask_for_students(self):
        t = Table("Name", "Matrikel", "Prüfung", title="Gefundene Studis")
        for studi in self.students.values():
            t.add_row(studi.full_name, str(studi.id or ""), studi.exam)
        get_console().print(t)
        more = questionary.text(
            "Mehr Studis? (ein Name pro Zeile)", multiline=True
        ).ask()
        if more:
            for studi in more.splitlines():
                if studi in self.students:
                    logger.warning("%s haben wir schon ...")
                else:
                    names = studi.split()
                    self.students[studi] = ExamAssassment(
                        name=names[-1],
                        given_name=" ".join(names[:-1]),
                        id=None,
                        exam_id=0,
                    )

    def ask_for_tasks(self, tasks: str | None = None, exams: str | None = None):
        if self.tasks:
            task_default = " ".join(f"A{i}" for i in range(1, len(self.tasks) + 1))
            if tasks == "auto":
                task_abbrevs = task_default.split()
            elif tasks:
                task_abbrevs = tasks.split()
            else:
                task_prompt = "\n".join(
                    f"  {i: 2d}. {task}" for i, task in enumerate(self.tasks, start=1)
                )
                task_abbrevs = (
                    questionary.text(
                        "Übungsaufgaben: ",
                        default=task_default,
                        instruction="Aufgabenkürzel in der Reihenfolge der folgenden Aufgaben:\n"
                        + task_prompt,
                        validate=lambda t: len(t.split()) == len(self.tasks),  # pyright: ignore[reportUnknownMemberType, reportUnknownArgumentType, reportUnknownLambdaType]
                    )
                    .ask()
                    .split()
                )
            self.task_abbr = dict(sorted(zip(task_abbrevs, self.tasks)))

        if exams is not None:
            exam_tasks_ = exams
        else:
            exam_tasks_ = questionary.text(
                "Klausuraufgaben (Anzahl oder leerzeichengetrennte Kürzel): "
            ).ask()
        if exam_tasks_:
            try:
                n = int(exam_tasks_)
                self.exam_tasks = [f"K{i}" for i in range(1, n + 1)]
            except ValueError:
                self.exam_tasks = exam_tasks_.split()

    def prepare_exam_sheet(self):
        ws = self.gradesheet

        # Spalten. Wir halten dabei auch die Koordinaten gleich fest.
        id_headings = ["Name", "Vorname", "Matrikel", "Prüfung"]

        if self.tasks:
            task_headings = list(self.task_abbr) + ["Σ Aufgaben", "% Aufgaben"]
        else:
            task_headings = []

        if self.exam_tasks:
            exam_headings = self.exam_tasks + ["Σ Klausur", "% Klausur"]
        else:
            exam_headings = []

        total_headings = ["% Gesamt", "Note"]

        bm = Bookmarks(
            max=2,
            student1=3,
            studentz=2 + len(self.students),
            avg=2 + len(self.students) + 2,
            task1=len(id_headings) + 1,
            taskz=len(id_headings) + len(self.tasks),
            exam1=len(id_headings) + len(task_headings) + 1,
            examz=len(id_headings) + len(task_headings) + len(self.exam_tasks),
            tasktot=len(id_headings) + len(self.tasks) + 1,
            taskpct=len(id_headings) + len(self.tasks) + 2,
            examtot=len(id_headings) + len(task_headings) + len(self.exam_tasks) + 1,
            exampct=len(id_headings) + len(task_headings) + len(self.exam_tasks) + 2,
            allpct=len(id_headings) + len(task_headings) + len(exam_headings) + 1,
            allgrade=len(id_headings) + len(task_headings) + len(exam_headings) + 2,
        )

        # Erste Zeile: Überschriften
        headings = id_headings + task_headings + exam_headings + total_headings
        heading_font = Font(size=12, bold=True)
        for i, heading in enumerate(headings, start=1):
            cell = ws.cell(1, i, heading)
            cell.font = heading_font
            cell.number_format = "0.0"
        for i, (_abbr, taskname) in enumerate(self.task_abbr.items(), start=bm.task1):
            ws.cell(1, i).comment = Comment(taskname, "")

        # Zweite Zeile: Maximale Punkte
        if self.tasks:
            for i, (task_name, task) in enumerate(self.tasks.items(), start=bm.task1):
                max_grade = first(task.values()).max_grade
                if not max_grade:
                    max_grade = max(stud.grade for stud in task.values())
                cell = ws.cell(bm.max, i, max_grade)
                cell.number_format = "0.0"
            total_cell = ws.cell(
                bm.max, bm.tasktot, bm.expand("=SUM({task1_}{max}:{taskz_}{max})")
            )
            total_cell.number_format = "0.0"
        label_cell = ws.cell(bm.max, bm.task1 - 1, "Max. ⏵")
        label_cell.font = Font(bold=True)
        if self.exam_tasks:
            for i in range(bm.exam1, bm.examz + 1):
                cell = ws.cell(bm.max, i)
                cell.number_format = "0.0"
            total_cell = ws.cell(
                bm.max, bm.examtot, bm.expand("=SUM({exam1_}{max}:{examz_}{max})")
            )
            total_cell.number_format = "0.0"

        # Eine Zeile pro Studi
        for row, (_name, student) in enumerate(self.students.items(), start=3):
            bm["row"] = row
            # Identifizierendes, TODO: Styling
            ws.cell(row, 1, student.name)
            ws.cell(row, 2, student.given_name)
            ws.cell(row, 3, student.id)
            exam_cell = ws.cell(row, 4, student.exam_id)
            if student.exam:
                exam_cell.comment = Comment(student.exam, "HIS")

            # Aufgaben
            if self.tasks:
                # Einzelne Aufgaben:
                for col, (_abbr, task_name) in enumerate(
                    self.task_abbr.items(), start=bm.task1
                ):
                    task = self.tasks[task_name].get(student.full_name)
                    points = task.grade if task is not None else None
                    cell = ws.cell(row, col, points)
                    cell.number_format = "0.0"

                # danach Summe und Prozent:
                total_cell = ws.cell(
                    row, bm.tasktot, bm.expand("=SUM({task1_}{row}:{taskz_}{row})")
                )
                total_cell.number_format = "0.0"
                percent_cell = ws.cell(
                    row, bm.taskpct, bm.expand("={tasktot_}{row}/{tasktot_}${max}")
                )
                percent_cell.number_format = "#0.0%"

            # Klausur
            if self.exam_tasks:
                for col in range(bm.exam1, bm.examz + 1):
                    cell = cast(Cell, ws.cell(row, col))  # it is not merged
                    cell.number_format = "0.0"

                total_cell = ws.cell(
                    row, bm.examtot, bm.expand("=SUM({exam1_}{row}:{examz_}{row})")
                )
                total_cell.number_format = "0.0"
                total_cell.font = Font(bold=True)

                percent_cell = ws.cell(
                    row, bm.exampct, bm.expand("={examtot_}{row}/{examtot_}${max}")
                )
                percent_cell.number_format = "#0.0%"

            if self.tasks and self.exam_tasks:
                all_pct = ws.cell(
                    row,
                    bm.allpct,
                    bm.expand("=0.35*{taskpct_}{row} + 0.65*{exampct_}{row}"),
                )
                all_pct.number_format = "#0.0%"
                all_grade = ws.cell(
                    row,
                    bm.allgrade,
                    bm.expand(
                        "=LOOKUP({allpct_}{row},Notenverteilung!A$2:A$12,Notenverteilung!B$2:B$12)"
                    ),
                )
                all_grade.number_format = "0.0"

        # Conditional Formatting
        for col in chain(range(bm.task1, bm.taskz + 2), range(bm.exam1, bm.examz + 2)):
            bmc = bm.using(col=col)
            points_scale = ColorScaleRule(  # pyright: ignore[reportUnknownVariableType]
                "num",
                0,
                Color("ff7f7f"),
                "formula",
                bmc.expand("={col_}${max}/2"),
                Color("ffff7f"),
                end_type="formula",
                end_value=bmc.expand("={col_}${max}"),
                end_color="80ff80",
            )
            ws.conditional_formatting.add(  # pyright: ignore[reportUnknownMemberType]
                bmc.expand("{col_}{student1}:{col_}{studentz}"),
                points_scale,  # pyright: ignore[reportUnknownArgumentType]
            )

        ws.conditional_formatting.add(  # pyright: ignore[reportUnknownMemberType]
            bm.expand("{allgrade_}{student1}:{allgrade_}{studentz}"),
            FormulaRule(  # pyright: ignore[reportUnknownArgumentType]
                [bm.expand("{allpct_}{student1} < 0.5")],
                stopIfTrue=False,
                fill=PatternFill(
                    fill_type="solid", start_color="ff7f7f", end_color="ff7f7f"
                ),
            ),
        )
        ws.conditional_formatting.add(  # pyright: ignore[reportUnknownMemberType]
            bm.expand("{allgrade_}{student1}:{allgrade_}{studentz}"),
            FormulaRule(  # pyright: ignore[reportUnknownArgumentType]
                [bm.expand("{allpct_}{student1} >= 0.5")],
                stopIfTrue=False,
                fill=PatternFill(
                    fill_type="solid", start_color="f7fff7", end_color="f7fff7"
                ),
            ),
        )

        # Durchschnitt etc.
        for col in chain(range(bm.task1, bm.taskz + 1), range(bm.exam1, bm.examz + 1)):
            bmc = bm.using(col=col)
            cell = ws.cell(
                bm.avg,
                col,
                bmc.expand("=AVERAGE({col_}{student1}:{col_}{studentz}) / {col_}{max}"),
            )
            cell.number_format = "#0.0%"

    def _create_grade_mapping(self) -> Worksheet:
        """Prepares the (static) table mapping points to grades."""
        ws = cast(Worksheet, self.workbook.create_sheet("Notenverteilung"))
        note_lookup = [
            ["Untergrenze", "Note"],
            [0.0, 5.0],
            [0.50, 4.0],
            [0.55, 3.7],
            [0.60, 3.3],
            [0.65, 3.0],
            [0.70, 2.7],
            [0.75, 2.3],
            [0.80, 2.0],
            [0.85, 1.7],
            [0.90, 1.3],
            [0.95, 1.0],
        ]
        for row in note_lookup:
            ws.append(row)
        heading = NamedStyle("Heading", font=Font(name="Ubuntu", bold=True, size=12))
        for row in ws.iter_rows(1, 1, 1, 2):
            for cell in row:
                cell.style = heading
        for row in ws.iter_rows(2):
            row[0].number_format = "0%"
            row[0].font = Font(name="Ubuntu")
            row[1].number_format = "0.0"
            row[1].font = Font(name="Ubuntu", bold=True)
        return ws

    def parse_his_sheet(self, file: Path):
        """
        Parses a single HIS sheet. Returns a dictionary mapping column heading to value for each student.
        """
        try:
            wb = load_workbook(file)
            ws = wb.active
            assert ws is not None
            rows = ws.iter_rows()
            while next(rows)[0].value != "startHISsheet":
                pass
            headings = tuple(str(cell.value) for cell in next(rows))
            raw_records: list[tuple[_CellGetValue, ...]] = []
            while record := next(rows):
                if record[0].value == "endHISsheet":
                    break
                raw_records.append(tuple(cell.value for cell in record))
            for row in raw_records:
                record = dict(zip(headings, row))
                yield ExamAssassment(
                    str(record["Nachname"]),
                    str(record["Vorname"]),
                    _convert(record["Matrikelnummer"], int),
                    _convert(record["PrüfungsNr."], int),
                    str(record["Titel"]),
                )
        except StopIteration:
            raise InvalidForm(f"{file} is not a HIS table")


def _convert[T, D](
    value: Any,  # pyright: ignore[reportExplicitAny]
    type_: type[T],
    default: D | None = None,
) -> T | D:
    try:
        return type_(value)  # pyright: ignore[reportCallIssue]
    except ValueError:
        if default is None:
            return type_()
        else:
            return default


def parse_wuecampus_sheet(
    file: Path,
) -> dict[str, list[TaskAssessment]]:
    """
    Parses a WueCampus sheet in CSV form.

    There are actually two kinds of sheet:

    (1) a single task sheet ID with the fields
         "Vollständiger Name", E-Mail-Adresse, Status, Bewertung, Bestwertung,
         "Bewertung kann geändert werden", "Zuletzt geändert (Abgabe)",
         "Zuletzt geändert (Bewertung)", "Feedback als Kommentar"

    (2) the total sheet that can be exported from the _Bewertungen_ section of wuecampus.
        This file contains fields Vorname, Nachname, Matrikelnr., Institution, Studiengang, E-Mail-Adresse,
        then for each task a field starting with "Aufgabe: ", and then a few summary fields
    """
    with file.open(newline="") as csvfile:
        records = list(DictReader(csvfile))

    result = {}

    if "Bewertung" in records[0]:  # single task
        if file.stem == "Bewertung":
            task_name = file.parent.stem
        else:
            task_name = file.stem.split("-")[2]
        students = [
            TaskAssessment(
                full_name=record["Vollständiger Name"],
                grade=_convert(record["Bewertung"], float),
                max_grade=_convert(record.get("Bestwertung", 1), float),
            )
            for record in records
        ]
        result = {task_name: students}
    else:
        for field in records[0]:
            if field.startswith("Aufgabe: "):
                task_name = field[8:].strip()
                students = [
                    TaskAssessment(
                        name=record["Nachname"],
                        given_name=record["Vorname"],
                        id=_convert(record["Matrikelnr."], int),
                        grade=_convert(record[field], float),
                    )
                    for record in records
                ]
                result[task_name] = students
    # clear all empty grades
    result = {
        task_name: [stud for stud in students if stud.grade]
        for task_name, students in result.items()
    }
    return result


@app.command()
def create(
    sources: list[Path],
    output: Path,
    tasks: str | None = None,
    exams: str | None = None,
):
    """
    Create an XLSX table to record grades.

    Args:
        sources: Source files. Each file can either be a HIS excel sheet exported from WueStudy, or a CSV sheet with task grades.
        output: output file with a single table
        tasks: auto, total number, or abbreviations for tasks columns: Generate task columns in the correct order. If missing, you will be asked interactively.
        exam: optional number or space separated abbreviations of exam questions. If missing, you will be asked interactively.
    """
    logging.basicConfig(
        level=logging.DEBUG,
        handlers=[RichHandler(show_time=False)],
        format="%(message)s",
    )
    logging.captureWarnings(True)
    gradesheet = GradeSheet(sources)
    gradesheet.merge_student_lists()
    gradesheet.ask_for_students()
    gradesheet.ask_for_tasks(tasks, exams)
    gradesheet.prepare_exam_sheet()
    gradesheet.workbook.save(output)
